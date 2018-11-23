# -*- coding: utf-8 -*-
# @Time    : 2018/10/17 21:10
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : test_api.py
#引入ddt
import unittest
from openpyxl import load_workbook
from ddt import ddt,data
from API_Frame_Work_1019.common import project_path
from API_Frame_Work_1019.common.http_request import HttpRequest
from API_Frame_Work_1019.common.do_excel import DoExcel
from API_Frame_Work_1019.common.my_log import  MyLog
from API_Frame_Work_1019.common.do_mysql import DoMysql

#获取测试数据
test_data=DoExcel(project_path.test_cases_path,'register').get_data()
#全局变量
COOKIES=None

@ddt
class TestApi(unittest.TestCase):
    def setUp(self):
        self.t=DoExcel(project_path.test_cases_path,'register')#创建一个Excel操作实例
        self.logger=MyLog()#实例
        self.logger.info("开始测试了")

    @data(*test_data)
    def test_api(self,data_item):
        global COOKIES#声明是全局变量
        self.logger.info("正在运行第{0}条用例：{1}".format(data_item['id'],data_item['description']))
        self.logger.info("测试数据是:{0}".format(data_item['param']))
        #请求之前 是不是要判断下 是否有loan_id 去数据库
        if data_item['param'].find('${loan_id}')!=-1:
            wb=load_workbook(project_path.test_cases_path)
            sheet=wb['info']
            member_id=sheet.cell(4,2).value
            #怎么去改写代码 灵活处理呢？  1)麻烦一点  2）都可以写到配置文件文件
            sql='select max(id) from loan where memberid={0}'.format(member_id)
            loan_id=DoMysql().do_mysql(sql)[0][0]
            param=data_item['param'].replace('${loan_id}',str(loan_id))#数字转为字符串
            sheet.cell(5,2).value=loan_id#写回查询到的loan_id到excel
            #保存
            wb.save(project_path.test_cases_path)
        else:
            param=data_item['param']
        res=HttpRequest().http_request(data_item['url'],eval(param),data_item['HttpMethod'],COOKIES)#全局变量
        #可以获取到cookie吗？
        if res.cookies:#任何非空数据的布尔值都为True  cookies是一个类字典的格式
            COOKIES=res.cookies#如果cookies不为空 就替换全局变量的COOKIES 修改全局变量

        self.logger.info("测试结果是：{0}".format(res.json()))
        #加断言
        try:
            self.assertEqual(str(data_item['ExpectedResult']),str(res.json()['code']))#res.json()['code']
            TestResult='PASS'
        except AssertionError as e:#断言异常
            TestResult='Failed'
            self.logger.error("请求出错了，错误是{0}".format(e))
            raise e#对异常处理完毕之后 要记得抛出
        finally:#最终都要写回到Excel
            self.t.write_back(data_item['id']+1,res.json()['code'],TestResult)

    def tearDown(self):
        self.logger.info("结束这个用例的测试")
# -*- coding: utf-8 -*-
# @Time    : 2018/10/15 21:12
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py

#处理测试数据的模块
from openpyxl import  load_workbook
from API_Frame_Work_1019.common.read_config import ReadConfig
from API_Frame_Work_1019.common import project_path
from API_Frame_Work_1019.common.get_info_data import GetInfoData

class DoExcel:
    def __init__(self,file_name):
        self.file_name=file_name
        self.sheet_list=eval(ReadConfig().read_config(project_path.case_config_path,'CASECONFIG','sheet_list'))

    def get_data(self):#获取测试数据
        wb=load_workbook(self.file_name)#打开的工作簿

        test_data=[]#存储所有的测试数据
        for key in self.sheet_list:#获取字典里面的每个key值  register login
            sheet=wb[key]#定位某个表单 表单名与key名一致 register
            if self.sheet_list[key]=='all':#读取所有的用例
                for i in range(2,sheet.max_row+1):#sheet.max_row sheet.max_column
                    sub_data={}#存储每一行测试测试用例数据  字典的格式
                    sub_data['id']=sheet.cell(i,1).value#用例序号
                    sub_data['HttpMethod']=sheet.cell(i,2).value
                    sub_data['module']=sheet.cell(i,3).value
                    sub_data['description']=sheet.cell(i,4).value
                    sub_data['url']=sheet.cell(i,5).value
                    sub_data['sheet_name']=key#存储表单名
                    #字符串替换
                    if sheet.cell(i,6).value.find('${tel}')!=-1:#修改点  对这个里面tel去进行替换 13667670163
                        sub_data['param']=sheet.cell(i,6).value.replace('${tel}',str(GetInfoData.no_reg_tel))
                    elif sheet.cell(i,6).value.find('${tel_1}')!=-1:
                        GetInfoData.no_reg_tel+=1#每次自加 自己跟代码就约定就好了
                        sub_data['param']=sheet.cell(i,6).value.replace('${tel_1}',str(GetInfoData.no_reg_tel))
                    elif sheet.cell(i,6).value.find('${login_tel}')!=-1:
                        sub_data['param']=sheet.cell(i,6).value.replace('${login_tel}',str(GetInfoData.login_tel))
                    elif sheet.cell(i,6).value.find('${loan_tel}')!=-1:
                        sub_data['param']=sheet.cell(i,6).value.replace('${loan_tel}',str(GetInfoData.loan_tel))
                    elif sheet.cell(i,6).value.find('${member_id}')!=-1:
                        sub_data['param']=sheet.cell(i,6).value.replace('${member_id}',str(GetInfoData.member_id))
                    elif sheet.cell(i,6).value.find('${invest_member_id}')!=-1:
                        sub_data['param']=sheet.cell(i,6).value.replace('${invest_member_id}',str(GetInfoData.invest_member_id))
                    else:#没有找到 不需要做替换的
                        sub_data['param']=sheet.cell(i,6).value
                    sub_data['ExpectedResult']=sheet.cell(i,7).value
                    test_data.append(sub_data)#完成一行的读取  存到sub_data 字典 然后append到test_data
            else:
                for i in self.sheet_list[key]:#sheet.max_row sheet.max_column
                    sub_data={}
                    #拓展点：把数据的获取 写成一个函数
                    sub_data['id']=sheet.cell(i+1,1).value#用例序号
                    sub_data['HttpMethod']=sheet.cell(i+1,2).value
                    sub_data['module']=sheet.cell(i+1,3).value
                    sub_data['description']=sheet.cell(i+1,4).value
                    sub_data['url']=sheet.cell(i+1,5).value
                    sub_data['sheet_name']=key#存储表单名
                    #字符串替换
                    if sheet.cell(i+1,6).value.find('${tel}')!=-1:#修改点  对这个里面tel去进行替换
                        sub_data['param']=sheet.cell(i+1,6).value.replace('${tel}',str(GetInfoData.no_reg_tel))
                    elif sheet.cell(i+1,6).value.find('${tel_1}')!=-1:
                        GetInfoData.no_reg_tel+=1#每次自加
                        sub_data['param']=sheet.cell(i+1,6).value.replace('${tel_1}',str(GetInfoData.no_reg_tel))
                    elif sheet.cell(i+1,6).value.find('${login_tel}')!=-1:
                        sub_data['param']=sheet.cell(i+1,6).value.replace('${login_tel}',str(GetInfoData.login_tel))
                    elif sheet.cell(i+1,6).value.find('${loan_tel}')!=-1:
                        sub_data['param']=sheet.cell(i+1,6).value.replace('${loan_tel}',str(GetInfoData.loan_tel))
                    elif sheet.cell(i+1,6).value.find('${member_id}')!=-1:
                        sub_data['param']=sheet.cell(i+1,6).value.replace('${member_id}',str(GetInfoData.member_id))
                    elif sheet.cell(i+1,6).value.find('${invest_member_id}')!=-1:
                        sub_data['param']=sheet.cell(i+1,6).value.replace('${invest_member_id}',str(GetInfoData.invest_member_id))
                    else:#没有找到 不需要做替换的
                        sub_data['param']=sheet.cell(i+1,6).value
                    sub_data['ExpectedResult']=sheet.cell(i+1,7).value
                    test_data.append(sub_data)
        #1022更改的点：用一个条件控制 是否要做更新
            if key=='register':#只有注册模块的时候才会去更新
                 wb['info'].cell(1,2).value=GetInfoData.no_reg_tel+1#未替换的手机写回到Excel
            wb.save(self.file_name)
        return test_data

    def write_back(self,sheet_name,row,ActualResult,TestResult):#写回数据到Excel
        wb=load_workbook(self.file_name)
        sheet=wb[sheet_name]
        sheet.cell(row,8).value=ActualResult
        sheet.cell(row,9).value=TestResult

        #写完之后要保存
        wb.save(self.file_name)

#测试代码
if __name__ == '__main__':
    test_data=DoExcel(project_path.test_cases_path).get_data()
    print(test_data)
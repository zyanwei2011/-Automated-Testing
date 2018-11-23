# -*- coding: utf-8 -*-
# @Time    : 2018/10/26 21:07
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : get_info_data.py
from openpyxl import load_workbook
from API_Frame_Work_1019.common import project_path
class GetInfoData:
    '''用来获取Excel里面info表单里面的数据'''
    wb=load_workbook(project_path.test_cases_path)

    #获取到tel表单 然后获取到最新未注册的手机号码
    no_reg_tel=wb['info'].cell(1,2).value#未注册的手机号

    #获取到登录的手机号 存储在info这个表单里面
    login_tel=wb['info'].cell(2,2).value

    #获取到专门用来加标的账户手机号
    loan_tel=wb['info'].cell(3,2).value

    #获取到加标的member_id
    member_id=wb['info'].cell(4,2).value

    #获取到投标的invest_member_id
    invest_member_id=wb['info'].cell(5,2).value

    #因为loan_id是没有的 所以我们这样做
    loan_id=None#定义一个初始值

    #COOKIES值
    COOKIES=None#初始值

if __name__ == '__main__':
   print(GetInfoData.no_reg_tel)
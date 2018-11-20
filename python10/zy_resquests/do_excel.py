import openpyxl


class Do_Excel:
    def read_data(self, file_path, sheet_name):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        test_data = []
        for i in range(2,sheet.max_row+1):
            sub_data= {}
            sub_data['id'] = sheet.cell(i, 1).value
            sub_data['module'] = sheet.cell(i, 2).value
            sub_data['title'] = sheet.cell(i, 3).value
            sub_data['method'] = sheet.cell(i, 4).value
            sub_data['param'] = sheet.cell(i, 5).value
            sub_data['ExpectedResult(code)'] = sheet.cell(i, 6).value
            test_data.append(sub_data)
        return test_data

    def write_back(self, file_path, sheet_name, row, ActualResult, TestResult):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        sheet.cell(row, 7).value = ActualResult
        sheet.cell(row, 8).value = TestResult
        wb.save(file_path)


if __name__ == '__main__':
    r = Do_Excel().read_data('test_data.xlsx', 'register')
    print(r)

import openpyxl


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_data(self,  button, case_id_list):
        wb = openpyxl.load_workbook(self.file_name)
        sheet_register = wb[self.sheet_name]
        sheet_tel = wb['info']
        no_reg_tel = sheet_tel.cell(1,2).value
        no_reg_tel_1 = no_reg_tel + 1
        no_reg_tel_2 = no_reg_tel + 2
        no_reg_tel_3 = no_reg_tel + 3
        no_reg_tel_4 = no_reg_tel + 4
        no_reg_tel_5 = no_reg_tel + 5

        test_data = []
        for i in range(2,sheet_register.max_row+1):
            sub_data= {}
            sub_data['id'] = sheet_register.cell(i, 1).value
            sub_data['module'] = sheet_register.cell(i, 2).value
            sub_data['title'] = sheet_register.cell(i, 3).value
            sub_data['method'] = sheet_register.cell(i, 4).value
            sub_data['url'] = sheet_register.cell(i, 5).value
            if sheet_register.cell(i, 6).value.find('${tel}') != -1:
                sub_data['param'] = sheet_register.cell(i, 6).value.replace('${tel}', str(no_reg_tel))
            elif sheet_register.cell(i, 6).value.find('${tel_1}') != -1:
                sub_data['param'] = sheet_register.cell(i, 6).value.replace('${tel_1}', str(no_reg_tel_1))
            elif sheet_register.cell(i, 6).value.find('${tel_2}') != -1:
                sub_data['param'] = sheet_register.cell(i, 6).value.replace('${tel_2}', str(no_reg_tel_2))
            elif sheet_register.cell(i, 6).value.find('${tel_3}') != -1:
                sub_data['param'] = sheet_register.cell(i, 6).value.replace('${tel_3}', str(no_reg_tel_3))
            elif sheet_register.cell(i, 6).value.find('${tel_4}') != -1:
                sub_data['param'] = sheet_register.cell(i, 6).value.replace('${tel_4}', str(no_reg_tel_4))
            else:
                sub_data['param'] = sheet_register.cell(i, 6).value
            sub_data['ExpectedResult(code)'] = sheet_register.cell(i, 7).value
            test_data.append(sub_data)

        if button == 'on':
            final_data = []
            for item in test_data:
                if item['id'] in eval(case_id_list):
                    final_data.append(item)
        else:
            final_data = test_data

        wb['info'].cell(1,2).value = no_reg_tel_5
        wb.save(self.file_name)
        return final_data

    def write_back(self, row, ActualResult, TestResult):
        wb = openpyxl.load_workbook(self.file_name)
        sheet = wb['register']
        sheet.cell(row, 8).value = ActualResult
        sheet.cell(row, 9).value = TestResult
        wb.save(self.file_name)


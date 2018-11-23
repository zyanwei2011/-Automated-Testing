# -*- coding: utf-8 -*-
# @Time    : 2018/10/15 21:12
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py

#处理测试数据的模块
from openpyxl import  load_workbook

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name#Excel
        self.sheet_name=sheet_name#表单名

    def get_data(self,button,case_id_list):#获取测试数据
        wb=load_workbook(self.file_name)#打开的工作簿
        sheet=wb[self.sheet_name]#1022修改点

        #获取到tel表单 然后获取到最新未注册的手机号码
        no_reg_tel=wb['info'].cell(1,2).value#未注册的手机号
        no_reg_tel_1=no_reg_tel+1#第二个未注册的手机号
        no_reg_tel_2=no_reg_tel+2#第三个未注册的手机号
        no_reg_tel_3=no_reg_tel+3#第四个未注册的手机号

        #获取到登录的手机号 存储在info这个表单里面
        login_tel=wb['info'].cell(2,2).value

        #获取到专门用来加标的账户手机号
        loan_tel=wb['info'].cell(3,2).value

        #获取到加标的member_id
        member_id=wb['info'].cell(4,2).value

        test_data=[]
        for i in range(2,sheet.max_row+1):#sheet.max_row sheet.max_column
            sub_data={}
            sub_data['id']=sheet.cell(i,1).value#用例序号
            sub_data['HttpMethod']=sheet.cell(i,2).value
            sub_data['module']=sheet.cell(i,3).value
            sub_data['description']=sheet.cell(i,4).value
            sub_data['url']=sheet.cell(i,5).value
            #字符串替换
            if sheet.cell(i,6).value.find('${tel}')!=-1:#修改点  对这个里面tel去进行替换
                sub_data['param']=sheet.cell(i,6).value.replace('${tel}',str(no_reg_tel))
            elif sheet.cell(i,6).value.find('${tel_1}')!=-1:
                sub_data['param']=sheet.cell(i,6).value.replace('${tel_1}',str(no_reg_tel_1))
            elif sheet.cell(i,6).value.find('${tel_2}')!=-1:
                sub_data['param']=sheet.cell(i,6).value.replace('${tel_2}',str(no_reg_tel_2))
            elif sheet.cell(i,6).value.find('${login_tel}')!=-1:
                sub_data['param']=sheet.cell(i,6).value.replace('${login_tel}',str(login_tel))
            elif sheet.cell(i,6).value.find('${loan_tel}')!=-1:
                sub_data['param']=sheet.cell(i,6).value.replace('${loan_tel}',str(loan_tel))
            elif sheet.cell(i,6).value.find('${member_id}')!=-1:
                sub_data['param']=sheet.cell(i,6).value.replace('${member_id}',str(member_id))
            else:#没有找到 不需要做替换的
                sub_data['param']=sheet.cell(i,6).value
            # #字典：eval转成字典--根据key 获取到mobilephone去比对是等于${tel}
            # param=eval(sheet.cell(i,5).value)
            # if param['mobilephone']=='${tel}':
            #     param['mobilephone']=str(no_reg_tel)
            #     sub_data['param']=str(param)
            # else:
            #     sub_data['param']=str(param)
            sub_data['ExpectedResult']=sheet.cell(i,7).value
            test_data.append(sub_data)
        #1022更改的点：用一个条件控制 是否要做更新
        if self.sheet_name=='register':#只有注册模块的时候才会去更新
             wb['info'].cell(1,2).value=no_reg_tel_3#未替换的手机写回到Excel
        wb.save(self.file_name)
        #可以根据case.config决定跑哪些用例
        if button=='on':#约定button=on的时候  就执行所有的用例
            finally_data=test_data
        else:#默认执行case_id_list里面的用例
            finally_data=[]
            for item in test_data:
                if item['id'] in eval(case_id_list):
                        finally_data.append(item)
        return finally_data

    def write_back(self,row,ActualResult,TestResult):#写回数据到Excel
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        sheet.cell(row,8).value=ActualResult
        sheet.cell(row,9).value=TestResult

        #写完之后要保存
        wb.save(self.file_name)

#测试代码
if __name__ == '__main__':
    #
    from API_Frame_Work_1019.common import project_path
    test_data=DoExcel(project_path.test_cases_path,"register").get_data("on",[1,2,3])
    print(test_data)
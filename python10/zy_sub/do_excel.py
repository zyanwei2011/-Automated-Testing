import openpyxl

class DoExcel:
    def read_data(self, button, case_id):
        wb = openpyxl.load_workbook('test_data.xlsx')
        sheet = wb['model1']
        header = []
        for column in range(1, 6):
            header.append(sheet.cell(1, column).value)
        # ['case_id', 'title', 'param_a', 'param_b', 'ExpectedResult', 'ActualResult', 'TestResult']

        test_data = []
        for row_id  in range(2, 9):
            row_data = {}
            for column_id in range(1, 6):
                row_data[header[column_id-1]] = sheet.cell(row_id, column_id).value
            test_data.append(row_data)

        # ## 根据配置文件过滤测试数据
        # button = ReadConfig().read_config('case.conf', 'SECTION', 'button')
        # case_id = ReadConfig().read_config('case.conf', 'SECTION', 'case_id')

        if button == 'on':       # 如果button为on时，执行配置文件中配置的用例，否则执行左右用例
            final_data = []
            for item in test_data:
                if item['case_id'] in case_id:
                    final_data.append(item)
        else:
            final_data = test_data
        return final_data

    def write_back(self, row_id, ActualResult, TestResult):
        wb = openpyxl.load_workbook('test_data.xlsx')
        sheet = wb['model1']
        sheet.cell(row_id, 6).value = ActualResult
        sheet.cell(row_id, 7).value = TestResult
        wb.save('test_data.xlsx')


# if __name__ == '__main__':
#     final_data = DoExcel().read_data('on', [1, 3, 5])
#     print(final_data)







# -*- coding: utf-8 -*-
# @Time    : 2018/10/15 21:12
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py

#处理测试数据的模块
from openpyxl import  load_workbook

class DoExcel:
    def get_data(self,file_name,sheet_name):#获取测试数据
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        #获取到tel表单 然后获取到最新未注册的手机号码
        # print(sheet.max_row)
        # print(sheet.max_column)
        test_data=[]
        for i in range(2,sheet.max_row+1):#sheet.max_row sheet.max_column
            sub_data={}
            sub_data['id']=sheet.cell(i,1).value#用例序号
            sub_data['HttpMethod']=sheet.cell(i,2).value
            sub_data['module']=sheet.cell(i,3).value
            sub_data['description']=sheet.cell(i,4).value
            sub_data['param']=sheet.cell(i,5).value#修改点  对这个里面tel去进行替换
            sub_data['ExpectedResult']=sheet.cell(i,6).value
            test_data.append(sub_data)
        #update tel 用完了之后 把取到的手机号+1  存回excel
        return test_data

    def update_tel(self,new_tel):#更新手机号码
        pass

    def write_back(self,file_name,sheet_name,row,ActualResult,TestResult):#写回数据到Excel
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]

        sheet.cell(row,7).value=ActualResult
        sheet.cell(row,8).value=TestResult

        #写完之后要保存
        wb.save(file_name)

if __name__ == '__main__':
    test_data=DoExcel().get_data("api.xlsx","testcase")
    print(test_data)
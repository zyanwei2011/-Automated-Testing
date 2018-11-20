# -*- coding: utf-8 -*-
# @Time    : 2018/10/15 21:12
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py

#处理测试数据的模块
from openpyxl import  load_workbook

class DoExcel:
    def __init__(self,file_name):
        self.file_name=file_name

    def get_data(self,button,case_id_list):#获取测试数据
        wb=load_workbook(self.file_name)#打开的工作簿
        sheet=wb['testcase']#这个是测试用例的表单

        #获取到tel表单 然后获取到最新未注册的手机号码
        no_reg_tel=wb['tel'].cell(1,1).value#整数

        test_data=[]
        for i in range(2,sheet.max_row+1):#sheet.max_row sheet.max_column
            sub_data={}
            sub_data['id']=sheet.cell(i,1).value#用例序号
            sub_data['HttpMethod']=sheet.cell(i,2).value
            sub_data['module']=sheet.cell(i,3).value
            sub_data['description']=sheet.cell(i,4).value
            sub_data['url']=sheet.cell(i,5).value
            #字符串替换
            if sheet.cell(i,6).value.find('${tel}')!=-1:#修改点  对这个里面tel去进行替换
                sub_data['param']=sheet.cell(i,6).value.replace('${tel}',str(no_reg_tel))
            else:#没有找到 不需要做替换的
                sub_data['param']=sheet.cell(i,6).value
            # #字典：eval转成字典--根据key 获取到mobilephone去比对是等于${tel}
            # param=eval(sheet.cell(i,5).value)
            # if param['mobilephone']=='${tel}':
            #     param['mobilephone']=str(no_reg_tel)
            #     sub_data['param']=str(param)
            # else:
            #     sub_data['param']=str(param)
            sub_data['ExpectedResult']=sheet.cell(i,7).value
            test_data.append(sub_data)
        #update tel 用完了之后 把取到的手机号+1  存回excel
        wb['tel'].cell(1,1).value=no_reg_tel+1
        wb.save(self.file_name)
        #可以根据case.config决定跑哪些用例
        if button=='on':#约定button=on的时候  就执行所有的用例
            finally_data=test_data
        else:#默认执行case_id_list里面的用例
            finally_data=[]
            for item in test_data:
                if item['id'] in eval(case_id_list):
                        finally_data.append(item)
        return finally_data

    def write_back(self,row,ActualResult,TestResult):#写回数据到Excel
        wb=load_workbook(self.file_name)
        sheet=wb['testcase']
        sheet.cell(row,8).value=ActualResult
        sheet.cell(row,9).value=TestResult

        #写完之后要保存
        wb.save(self.file_name)

if __name__ == '__main__':
    test_data=DoExcel().get_data("api.xlsx","testcase")
    print(test_data)
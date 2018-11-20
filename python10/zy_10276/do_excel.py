import openpyxl
# from zy_10276.read_config import ReadConfig


class DoExcel:
    def __init__(self, file_name):
        self.file_name = file_name

    def read_data(self,  button, case_id_list):
        wb = openpyxl.load_workbook(self.file_name)
        sheet = wb['register']

        test_data = []
        for i in range(2,sheet.max_row+1):
            sub_data= {}
            sub_data['id'] = sheet.cell(i, 1).value
            sub_data['module'] = sheet.cell(i, 2).value
            sub_data['title'] = sheet.cell(i, 3).value
            sub_data['method'] = sheet.cell(i, 4).value
            sub_data['param'] = sheet.cell(i, 5).value
            sub_data['ExpectedResult(code)'] = sheet.cell(i, 6).value
            test_data.append(sub_data)

        if button == 'on':
            final_data = []
            for item in test_data:
                if item['id'] in eval(case_id_list):
                    final_data.append(item)
        else:
            final_data = test_data
        return final_data

    def write_back(self, row, ActualResult, TestResult):
        wb = openpyxl.load_workbook(self.file_name)
        sheet = wb['register']
        sheet.cell(row, 7).value = ActualResult
        sheet.cell(row, 8).value = TestResult
        wb.save(self.file_name)


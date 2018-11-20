from openpyxl import load_workbook

wb=load_workbook("test_data.xlsx")
sheet=wb['test_data']

# sheet.cell(row,6).value=0
# sheet.cell(row,7).value='PASS'

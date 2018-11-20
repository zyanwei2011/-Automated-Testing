# -*- coding: utf-8 -*-
# @Time    : 2018/9/21 20:42
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py
from openpyxl import  load_workbook

#第一个版本  列表嵌套列表
# class DoExcel:
#     def do_excel(self):
#         wb=load_workbook("test_data.xlsx")
#         sheet=wb['test_data']
#         #把每一行的数据读取出来# 2 3 4 5 6 7 range(2,8)
#         test_data=[]
#         for i in range(2,8):#i=2 i=3
#             sub_data=[]
#             for j in range(1,6):#j= 1 2 3 4 5
#                sub_data.append(sheet.cell(i,j).value)
#             test_data.append(sub_data)
#         return test_data#返回测试数据

#第二种方法：列表嵌套字典（列表里面的元素都是字典类型）
class DoExcel:
    def do_excel(self):
        wb=load_workbook("test_data.xlsx")
        sheet=wb['test_data']

        #先读取表头，也就是第一行：
        #header=['id','title','Param_a','Param_b','ExpectedResult']
        header=[]
        for i in range(1,6):
            header.append(sheet.cell(1,i).value)

        test_data=[]
        for i in range(2,8):#i=2
            sub_data={}
            for j in range(1,6):#j=1 2 3 4 5  2,1
               sub_data[header[j-1]]=sheet.cell(i,j).value
            test_data.append(sub_data)
        return test_data#返回测试数据

    #写一个函数：写回结果到Excel  每一行的第六列 第七列
    def write_back(self,row,ActualResult,TestResult):
        wb=load_workbook("test_data.xlsx")
        sheet=wb['test_data']
        #写入实际的结果值  列是固定的
        sheet.cell(row,6).value=ActualResult#运算结果的实际值

        #写入测试结果是通过还是不通过 列是固定的
        sheet.cell(row,7).value=TestResult

        #一定要保存
        wb.save("test_data.xlsx")


#第三种方法：最简单 必须掌握
# class DoExcel:
#     def do_excel(self):
#         wb=load_workbook("test_data.xlsx")
#         sheet=wb['test_data']
#
#         test_data=[]
#         for i in range(2,8):#只对行去进行循环
#             sub_data={}
#             sub_data['id']=sheet.cell(i,1).value
#             sub_data['title']=sheet.cell(i,2).value
#             sub_data['Param_a']=sheet.cell(i,3).value
#             sub_data['Param_b']=sheet.cell(i,4).value
#             sub_data['ExpectedResult']=sheet.cell(i,5).value
#             test_data.append(sub_data)
#         return test_data#返回测试数据

if __name__ == '__main__':
   test_data=DoExcel().do_excel()
   print("获取到的测试是：{0}".format(test_data))
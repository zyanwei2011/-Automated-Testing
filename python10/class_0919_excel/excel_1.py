# -*- coding: utf-8 -*-
# @Time    : 2018/9/19 22:12
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : excel_1.py
#excel openpyxl pip install openpyxl
#文件 .xlsx  一定要自己新建
#打开工作簿（excel)--定位表单（sheet)--定位单元格
#读取值

from openpyxl import load_workbook

#打开Excel
wb=load_workbook("python10.xlsx")#返回的是工作簿

#定位表单
# sheet=wb.get_sheet_by_name("name")
sheet=wb['name']

#定位单元格并取值
result=sheet.cell(2,2).value
print(result)

#修改值
sheet.cell(2,2).value="yulan"#赋值语句
#一定要保存 否则修改不生效
wb.save("python10.xlsx")

#疑问：请问 跟我讲的单元测试有啥关系？
#2：如果有关系 怎么去设计参数呢？
#3：设计好了 怎么读出来了呢？
